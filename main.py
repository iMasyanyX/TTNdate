import openpyxl
import os
from tkinter import Tk, Label, Button, messagebox
from tkinter import ttk
from tkcalendar import Calendar
from tkinter.filedialog import askdirectory

def select_directory():
    global directory_path
    directory_path = askdirectory()

def select_date():
    def set_date():
        selected_date = cal.selection_get()
        set_date_in_excel(selected_date)

    def set_date_in_excel(selected_date):
        for filename in os.listdir(directory_path):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(directory_path, filename)

                workbook = openpyxl.load_workbook(file_path)

                sheet_names = workbook.sheetnames
                sheet = workbook[sheet_names[0]]

                # Преобразование формата даты для openpyxl
                formatted_date = selected_date.strftime('%d.%m.%Y')
                date_day = selected_date.strftime('%d')
                date_month = selected_date.strftime('%m')
                date_year = selected_date.strftime('%Y')

                def month_name(month_number):
                    months = {
                        1: "января",
                        2: "февраля",
                        3: "марта",
                        4: "апреля",
                        5: "мая",
                        6: "июня",
                        7: "июля",
                        8: "августа",
                        9: "сентября",
                        10: "октября",
                        11: "ноября",
                        12: "декабря"
                    }
                    return months.get(month_number, "")

                month = month_name(
                    int(date_month))  # Преобразуем номер месяца в целое число перед вызовом функции month_name()
                formatted_month = date_day + " " + month + " " + date_year

                # Числовое представление месяца
                sheet.cell(row=row_number_numeric, column=column_number_numeric).value = formatted_date

                # Прописью месяц
                sheet.cell(row=row_number_word, column=column_number_word).value = formatted_month

                workbook.save(file_path)

        messagebox.showinfo('Успех', 'Даты успешно установлены во всех ТТН.')

    top = Tk()
    top.title("ТТН by MasyanyX")

    label_directory = Label(top, text="Выберите папку с файлами ТТН в формате .xlsx")
    label_directory.pack(pady=10)

    button_directory = Button(top, text="Выбрать папку", command=select_directory)
    button_directory.pack(pady=5)

    cal = Calendar(top, date_pattern="dd.MM.yyyy", selectmode="day")
    cal.pack(padx=10, pady=10)

    button_frame = ttk.Frame(top)
    button_frame.pack(pady=10)

    ok_button = ttk.Button(button_frame, text="Выполнить", command=set_date)
    ok_button.grid(row=0, column=0, padx=5)

    cancel_button = ttk.Button(button_frame, text="Закрыть", command=top.destroy)
    cancel_button.grid(row=0, column=1, padx=5)

    top.mainloop()

if __name__ == '__main__':
    directory_path = ""
    row_number_numeric = 5
    column_number_numeric = 19
    row_number_word = 60
    column_number_word = 3